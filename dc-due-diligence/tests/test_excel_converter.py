"""Tests for the Excel file conversion converter."""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from converters import ConfidenceLevel, ExtractionResult
from converters.excel import (
    ExcelConverter,
    _count_data,
    _format_cell,
    _rows_to_markdown_table,
    _trim_empty_rows,
)


# ------------------------------------------------------------------
# Fixtures
# ------------------------------------------------------------------

OPPORTUNITY_DIR = Path(__file__).resolve().parent.parent / "opportunity-example"


@pytest.fixture
def converter():
    return ExcelConverter()


@pytest.fixture
def simple_xlsx(tmp_path):
    """Create a simple .xlsx file with one sheet of data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Site", "Power (MW)", "Cost ($/kWh)"])
    ws.append(["Dallas", 10, 0.045])
    ws.append(["Phoenix", 25, 0.038])
    ws.append(["Chicago", 15, 0.052])
    path = tmp_path / "test_data.xlsx"
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def multi_sheet_xlsx(tmp_path):
    """Create a .xlsx file with multiple sheets, including an empty one."""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Financials"
    ws1.append(["Year", "Revenue", "EBITDA"])
    ws1.append([2024, 5000000, 1200000])
    ws1.append([2025, 7500000, 2100000])

    ws2 = wb.create_sheet("Specifications")
    ws2.append(["Attribute", "Value"])
    ws2.append(["Total Power", "50 MW"])
    ws2.append(["Floor Space", "120,000 sqft"])
    ws2.append(["Cooling", "Chilled Water"])

    # Empty sheet -- should be skipped.
    wb.create_sheet("Notes")

    path = tmp_path / "multi_sheet.xlsx"
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def xlsx_with_formulas(tmp_path):
    """Create a .xlsx file with formulas (saved without cached values).

    When opened with data_only=True, formula cells with no cached value
    return None.  The converter should handle this gracefully.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calculations"
    ws["A1"] = "Price"
    ws["B1"] = "Quantity"
    ws["C1"] = "Total"
    ws["A2"] = 100
    ws["B2"] = 5
    ws["C2"] = "=A2*B2"  # Formula -- cached value depends on Excel saving
    path = tmp_path / "formulas.xlsx"
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def xlsx_with_special_chars(tmp_path):
    """Create a .xlsx with pipe chars and newlines in cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Special"
    ws.append(["Name", "Description"])
    ws.append(["Test|Item", "Line1\nLine2"])
    ws.append(["Normal", "Simple text"])
    path = tmp_path / "special.xlsx"
    wb.save(str(path))
    wb.close()
    return path


# ------------------------------------------------------------------
# Unit tests for helper functions
# ------------------------------------------------------------------


class TestFormatCell:
    def test_none_becomes_empty(self):
        assert _format_cell(None) == ""

    def test_string_preserved(self):
        assert _format_cell("Hello World") == "Hello World"

    def test_string_stripped(self):
        assert _format_cell("  padded  ") == "padded"

    def test_integer(self):
        assert _format_cell(42) == "42"

    def test_float_whole_number(self):
        assert _format_cell(10.0) == "10"

    def test_float_decimal(self):
        assert _format_cell(3.14) == "3.14"

    def test_float_trailing_zeros(self):
        # 0.500000 should become "0.5"
        assert _format_cell(0.5) == "0.5"

    def test_boolean_true(self):
        assert _format_cell(True) == "Yes"

    def test_boolean_false(self):
        assert _format_cell(False) == "No"

    def test_newlines_replaced(self):
        assert _format_cell("line1\nline2") == "line1 line2"

    def test_carriage_return_replaced(self):
        assert _format_cell("line1\rline2") == "line1 line2"

    def test_zero(self):
        assert _format_cell(0) == "0"

    def test_zero_float(self):
        assert _format_cell(0.0) == "0"

    def test_negative_float(self):
        assert _format_cell(-3.5) == "-3.5"


class TestCountData:
    def test_empty_rows(self):
        rows: list[list[str]] = [["", ""], ["", ""]]
        data_rows, cells = _count_data(rows)
        assert data_rows == 0
        assert cells == 0

    def test_mixed_rows(self):
        rows = [["A", "B"], ["", ""], ["C", ""]]
        data_rows, cells = _count_data(rows)
        assert data_rows == 2
        assert cells == 3

    def test_all_filled(self):
        rows = [["X", "Y"], ["1", "2"]]
        data_rows, cells = _count_data(rows)
        assert data_rows == 2
        assert cells == 4


class TestTrimEmptyRows:
    def test_no_trimming_needed(self):
        rows = [["A"], ["B"]]
        assert _trim_empty_rows(rows) == [["A"], ["B"]]

    def test_leading_empty(self):
        rows = [[""], [""], ["A"]]
        assert _trim_empty_rows(rows) == [["A"]]

    def test_trailing_empty(self):
        rows = [["A"], [""], [""]]
        assert _trim_empty_rows(rows) == [["A"]]

    def test_both_ends(self):
        rows = [[""], ["A"], ["B"], [""]]
        assert _trim_empty_rows(rows) == [["A"], ["B"]]

    def test_all_empty(self):
        rows = [[""], [""]]
        assert _trim_empty_rows(rows) == []

    def test_empty_input(self):
        assert _trim_empty_rows([]) == []


class TestRowsToMarkdownTable:
    def test_simple_table(self):
        rows = [["Name", "Value"], ["Power", "10MW"]]
        md = _rows_to_markdown_table(rows)
        lines = md.split("\n")
        assert lines[0] == "| Name | Value |"
        assert lines[1] == "| --- | --- |"
        assert lines[2] == "| Power | 10MW |"

    def test_uneven_rows_padded(self):
        rows = [["A", "B", "C"], ["x"]]
        md = _rows_to_markdown_table(rows)
        lines = md.split("\n")
        assert lines[2].count("|") == 4  # 3 columns + outer pipes

    def test_pipe_escaped(self):
        rows = [["Col"], ["a|b"]]
        md = _rows_to_markdown_table(rows)
        assert "a\\|b" in md

    def test_empty_rows_skipped(self):
        rows = [["H1", "H2"], ["", ""], ["A", "B"]]
        md = _rows_to_markdown_table(rows)
        lines = md.split("\n")
        # Header + separator + one data row (empty row skipped)
        assert len(lines) == 3

    def test_empty_input(self):
        assert _rows_to_markdown_table([]) == ""

    def test_all_empty_cells(self):
        rows = [["", ""], ["", ""]]
        # After trimming, there should be no data.
        assert _rows_to_markdown_table(rows) == ""

    def test_header_only(self):
        rows = [["A", "B"]]
        md = _rows_to_markdown_table(rows)
        lines = md.split("\n")
        assert len(lines) == 2  # Header + separator


# ------------------------------------------------------------------
# Converter behaviour tests
# ------------------------------------------------------------------


class TestExcelConverterBasics:
    def test_supported_extensions(self, converter):
        assert ".xlsx" in converter.supported_extensions
        assert ".xlsb" in converter.supported_extensions

    def test_can_handle_xlsx(self, converter):
        assert converter.can_handle(Path("data.xlsx")) is True
        assert converter.can_handle(Path("DATA.XLSX")) is True

    def test_can_handle_xlsb(self, converter):
        assert converter.can_handle(Path("model.xlsb")) is True
        assert converter.can_handle(Path("MODEL.XLSB")) is True

    def test_cannot_handle_non_excel(self, converter):
        assert converter.can_handle(Path("doc.pdf")) is False
        assert converter.can_handle(Path("doc.docx")) is False
        assert converter.can_handle(Path("doc.csv")) is False

    def test_missing_file_returns_error(self, converter):
        result = converter.convert(Path("/tmp/nonexistent_excel_12345.xlsx"))
        assert result.success is False
        assert result.confidence == ConfidenceLevel.LOW
        assert "not found" in result.error.lower()
        assert result.method == "openpyxl"

    def test_missing_xlsb_file_method(self, converter):
        result = converter.convert(Path("/tmp/nonexistent_12345.xlsb"))
        assert result.success is False
        assert result.method == "pyxlsb"

    def test_wrong_extension_returns_error(self, converter, tmp_path):
        txt = tmp_path / "test.txt"
        txt.write_text("not excel")
        result = converter.convert(txt)
        assert result.success is False
        assert "not an excel" in result.error.lower()

    def test_result_type(self, converter):
        result = converter.convert(Path("/tmp/nonexistent_12345.xlsx"))
        assert isinstance(result, ExtractionResult)


class TestXlsxExtraction:
    def test_simple_file(self, converter, simple_xlsx):
        result = converter.convert(simple_xlsx)
        assert result.success is True
        assert result.method == "openpyxl"
        assert result.confidence in (ConfidenceLevel.HIGH, ConfidenceLevel.MEDIUM)
        assert result.page_count == 1  # One sheet with data

    def test_contains_sheet_header(self, converter, simple_xlsx):
        result = converter.convert(simple_xlsx)
        assert "## Summary" in result.text

    def test_contains_data(self, converter, simple_xlsx):
        result = converter.convert(simple_xlsx)
        assert "Dallas" in result.text
        assert "Phoenix" in result.text
        assert "Chicago" in result.text

    def test_markdown_table_format(self, converter, simple_xlsx):
        result = converter.convert(simple_xlsx)
        assert "|" in result.text
        assert "---" in result.text
        # Header row should contain column names.
        assert "Site" in result.text
        assert "Power (MW)" in result.text

    def test_metadata(self, converter, simple_xlsx):
        result = converter.convert(simple_xlsx)
        assert result.metadata["total_sheets"] == 1
        assert result.metadata["sheets_extracted"] == 1
        assert result.metadata["sheets_skipped"] == 0
        assert result.metadata["total_data_rows"] > 0
        assert result.metadata["total_cells_with_data"] > 0

    def test_multi_sheet(self, converter, multi_sheet_xlsx):
        result = converter.convert(multi_sheet_xlsx)
        assert result.success is True
        assert "## Financials" in result.text
        assert "## Specifications" in result.text
        # Empty "Notes" sheet should be skipped.
        assert "## Notes" not in result.text
        assert result.metadata["total_sheets"] == 3
        assert result.metadata["sheets_extracted"] == 2
        assert result.metadata["sheets_skipped"] == 1
        assert result.page_count == 2

    def test_formulas_handled_gracefully(self, converter, xlsx_with_formulas):
        """Formulas without cached values should appear as empty cells, not crash."""
        result = converter.convert(xlsx_with_formulas)
        assert result.success is True
        assert "Price" in result.text
        assert "Quantity" in result.text

    def test_special_characters(self, converter, xlsx_with_special_chars):
        result = converter.convert(xlsx_with_special_chars)
        assert result.success is True
        # Pipe should be escaped.
        assert "Test\\|Item" in result.text
        # Newline should be replaced with space.
        assert "Line1 Line2" in result.text

    def test_source_path_resolved(self, converter, simple_xlsx):
        result = converter.convert(simple_xlsx)
        assert result.source_path == simple_xlsx.resolve()

    def test_confidence_high_for_substantial_data(self, converter, multi_sheet_xlsx):
        result = converter.convert(multi_sheet_xlsx)
        assert result.confidence == ConfidenceLevel.HIGH


class TestXlsxEdgeCases:
    def test_completely_empty_workbook(self, converter, tmp_path):
        """A workbook with no data should succeed with low confidence."""
        wb = openpyxl.Workbook()
        path = tmp_path / "empty.xlsx"
        wb.save(str(path))
        wb.close()

        result = converter.convert(path)
        assert result.success is True
        assert result.confidence == ConfidenceLevel.LOW
        assert result.page_count == 0

    def test_single_cell(self, converter, tmp_path):
        """A workbook with a single cell should still work."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Only data"
        path = tmp_path / "single_cell.xlsx"
        wb.save(str(path))
        wb.close()

        result = converter.convert(path)
        assert result.success is True
        assert "Only data" in result.text

    def test_wide_spreadsheet(self, converter, tmp_path):
        """A sheet with many columns should format correctly."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Wide"
        headers = [f"Col{i}" for i in range(20)]
        ws.append(headers)
        ws.append([str(i) for i in range(20)])
        path = tmp_path / "wide.xlsx"
        wb.save(str(path))
        wb.close()

        result = converter.convert(path)
        assert result.success is True
        assert "Col0" in result.text
        assert "Col19" in result.text

    def test_numeric_types(self, converter, tmp_path):
        """Various numeric types should be formatted cleanly."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Numbers"
        ws.append(["Type", "Value"])
        ws.append(["Integer", 42])
        ws.append(["Float", 3.14159])
        ws.append(["Whole Float", 100.0])
        ws.append(["Negative", -5.5])
        ws.append(["Zero", 0])
        path = tmp_path / "numbers.xlsx"
        wb.save(str(path))
        wb.close()

        result = converter.convert(path)
        assert result.success is True
        assert "42" in result.text
        assert "100" in result.text  # 100.0 displayed as 100
        assert "-5.5" in result.text
        assert "0" in result.text


# ------------------------------------------------------------------
# Integration tests against real opportunity Excel files
# ------------------------------------------------------------------


class TestRealExcelFiles:
    """Integration tests against actual example spreadsheets.

    Skipped if the opportunity-example folder is missing.
    """

    @pytest.fixture
    def converter(self):
        return ExcelConverter()

    def _get_file(self, name: str) -> Path:
        f = OPPORTUNITY_DIR / name
        if not f.exists():
            pytest.skip(f"Example file not found: {name}")
        return f

    def test_xlsx_carrier_matrix(self, converter):
        xlsx = self._get_file("0d. DN Carrier Matrix.xlsx")
        result = converter.convert(xlsx)

        assert result.success is True
        assert result.method == "openpyxl"
        assert result.page_count > 0
        assert len(result.text) > 50
        assert result.confidence in (ConfidenceLevel.HIGH, ConfidenceLevel.MEDIUM)

    def test_xlsb_pro_forma(self, converter):
        xlsb = self._get_file("17. DataNovaX Pioneer Park  Pro Forma.xlsb")
        result = converter.convert(xlsb)

        assert result.success is True
        assert result.method == "pyxlsb"
        assert result.page_count > 0
        assert len(result.text) > 50

    def test_xlsx_has_metadata(self, converter):
        xlsx = self._get_file("0d. DN Carrier Matrix.xlsx")
        result = converter.convert(xlsx)

        assert "total_sheets" in result.metadata
        assert "sheets_extracted" in result.metadata
        assert "total_data_rows" in result.metadata
        assert "total_cells_with_data" in result.metadata

    def test_xlsb_has_metadata(self, converter):
        xlsb = self._get_file("17. DataNovaX Pioneer Park  Pro Forma.xlsb")
        result = converter.convert(xlsb)

        assert "total_sheets" in result.metadata
        assert "sheets_extracted" in result.metadata
        assert "total_data_rows" in result.metadata

    def test_xlsx_contains_markdown_tables(self, converter):
        xlsx = self._get_file("0d. DN Carrier Matrix.xlsx")
        result = converter.convert(xlsx)

        assert "|" in result.text
        assert "---" in result.text

    def test_xlsb_contains_markdown_tables(self, converter):
        xlsb = self._get_file("17. DataNovaX Pioneer Park  Pro Forma.xlsb")
        result = converter.convert(xlsb)

        assert "|" in result.text
        assert "---" in result.text

    def test_all_example_excels_process_without_error(self, converter):
        """Every Excel file in the example folder should return a result."""
        if not OPPORTUNITY_DIR.exists():
            pytest.skip("Example folder not found")

        xlsx_files = list(OPPORTUNITY_DIR.glob("*.xlsx"))
        xlsb_files = list(OPPORTUNITY_DIR.glob("*.xlsb"))
        all_files = xlsx_files + xlsb_files

        if not all_files:
            pytest.skip("No Excel files found in example folder")

        for f in all_files:
            result = converter.convert(f)
            assert isinstance(result, ExtractionResult), f"Bad result for {f.name}"
            assert result.success is True, f"Failed on {f.name}: {result.error}"
            assert result.page_count > 0, f"No sheets extracted for {f.name}"
