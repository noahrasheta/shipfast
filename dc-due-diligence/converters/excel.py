"""
Excel file conversion for .xlsx and .xlsb formats.

Handles both standard XML-based (.xlsx) spreadsheets via openpyxl and binary
(.xlsb) spreadsheets via pyxlsb.  Each sheet in the workbook is extracted as
a separate section with the sheet name as a markdown heading, and row data is
formatted as a markdown table.

Uses ``data_only=True`` for .xlsx files so computed cell values are returned
instead of the underlying formulas.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pyxlsb

from converters.base import BaseConverter, ConfidenceLevel, ExtractionResult


class ExcelConverter(BaseConverter):
    """Extract text from Excel workbooks (.xlsx and .xlsb).

    For .xlsx files, ``openpyxl`` is used with ``data_only=True`` to read
    computed values rather than formulas.  For .xlsb binary files, ``pyxlsb``
    is used.

    All sheets are extracted and formatted as markdown sections with a heading
    per sheet and a markdown table for the data.  Empty sheets and sheets
    containing only formatting (no data) are skipped.
    """

    supported_extensions: list[str] = [".xlsx", ".xlsb"]

    def convert(self, path: Path) -> ExtractionResult:
        """Extract all sheets from an Excel workbook and return markdown."""
        path = Path(path).resolve()

        if not path.exists():
            return ExtractionResult(
                source_path=path,
                text="",
                method=self._method_for(path),
                success=False,
                confidence=ConfidenceLevel.LOW,
                confidence_reason="file not found",
                error=f"File not found: {path}",
            )

        suffix = path.suffix.lower()
        if suffix not in self.supported_extensions:
            return ExtractionResult(
                source_path=path,
                text="",
                method="openpyxl",
                success=False,
                confidence=ConfidenceLevel.LOW,
                confidence_reason="unsupported file type",
                error=f"Not an Excel file: {suffix}",
            )

        try:
            if suffix == ".xlsx":
                return self._extract_xlsx(path)
            else:
                return self._extract_xlsb(path)
        except Exception as exc:
            return ExtractionResult(
                source_path=path,
                text="",
                method=self._method_for(path),
                success=False,
                confidence=ConfidenceLevel.LOW,
                confidence_reason="extraction crashed",
                error=f"Extraction failed: {exc}",
            )

    # ------------------------------------------------------------------
    # Format-specific extraction
    # ------------------------------------------------------------------

    def _extract_xlsx(self, path: Path) -> ExtractionResult:
        """Extract sheets from a .xlsx workbook using openpyxl."""
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        try:
            sheet_names = wb.sheetnames
            sections: list[str] = []
            total_data_rows = 0
            total_cells_with_data = 0
            sheets_extracted = 0

            for name in sheet_names:
                ws = wb[name]
                rows = _read_openpyxl_sheet(ws)

                if not rows:
                    continue

                data_rows, cells_with_data = _count_data(rows)
                if data_rows == 0:
                    continue

                total_data_rows += data_rows
                total_cells_with_data += cells_with_data
                sheets_extracted += 1

                md_table = _rows_to_markdown_table(rows)
                sections.append(f"## {name}\n\n{md_table}")

            return self._build_result(
                path=path,
                method="openpyxl",
                sections=sections,
                total_sheets=len(sheet_names),
                sheets_extracted=sheets_extracted,
                total_data_rows=total_data_rows,
                total_cells_with_data=total_cells_with_data,
            )
        finally:
            wb.close()

    def _extract_xlsb(self, path: Path) -> ExtractionResult:
        """Extract sheets from a .xlsb workbook using pyxlsb."""
        with pyxlsb.open_workbook(str(path)) as wb:
            sheet_names = wb.sheets
            sections: list[str] = []
            total_data_rows = 0
            total_cells_with_data = 0
            sheets_extracted = 0

            for name in sheet_names:
                with wb.get_sheet(name) as sheet:
                    rows = _read_pyxlsb_sheet(sheet)

                if not rows:
                    continue

                data_rows, cells_with_data = _count_data(rows)
                if data_rows == 0:
                    continue

                total_data_rows += data_rows
                total_cells_with_data += cells_with_data
                sheets_extracted += 1

                md_table = _rows_to_markdown_table(rows)
                sections.append(f"## {name}\n\n{md_table}")

            return self._build_result(
                path=path,
                method="pyxlsb",
                sections=sections,
                total_sheets=len(sheet_names),
                sheets_extracted=sheets_extracted,
                total_data_rows=total_data_rows,
                total_cells_with_data=total_cells_with_data,
            )

    # ------------------------------------------------------------------
    # Result building
    # ------------------------------------------------------------------

    def _build_result(
        self,
        path: Path,
        method: str,
        sections: list[str],
        total_sheets: int,
        sheets_extracted: int,
        total_data_rows: int,
        total_cells_with_data: int,
    ) -> ExtractionResult:
        """Assemble the final ExtractionResult from extracted sheet sections."""
        combined_text = "\n\n".join(sections)
        total_chars = len(combined_text)

        if sheets_extracted == 0:
            confidence = ConfidenceLevel.LOW
            confidence_reason = (
                f"no data found across {total_sheets} sheet(s)"
            )
        elif total_cells_with_data < 5:
            confidence = ConfidenceLevel.LOW
            confidence_reason = (
                f"very little data, only {total_cells_with_data} cells "
                f"with content across {sheets_extracted} sheet(s)"
            )
        elif total_data_rows < 3:
            confidence = ConfidenceLevel.MEDIUM
            confidence_reason = (
                f"sparse data, {total_data_rows} rows across "
                f"{sheets_extracted} sheet(s)"
            )
        else:
            confidence = ConfidenceLevel.HIGH
            confidence_reason = (
                f"{total_data_rows} data rows across "
                f"{sheets_extracted} sheet(s)"
            )

        metadata: dict[str, Any] = {
            "total_sheets": total_sheets,
            "sheets_extracted": sheets_extracted,
            "sheets_skipped": total_sheets - sheets_extracted,
            "total_data_rows": total_data_rows,
            "total_cells_with_data": total_cells_with_data,
            "total_chars": total_chars,
        }

        return ExtractionResult(
            source_path=path,
            text=combined_text,
            method=method,
            success=True,
            confidence=confidence,
            confidence_reason=confidence_reason,
            page_count=sheets_extracted,
            metadata=metadata,
        )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _method_for(path: Path) -> str:
        """Return the method label based on the file extension."""
        if path.suffix.lower() == ".xlsb":
            return "pyxlsb"
        return "openpyxl"


# ------------------------------------------------------------------
# Module-level helpers
# ------------------------------------------------------------------


def _read_openpyxl_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> list[list[str]]:
    """Read all rows from an openpyxl worksheet, returning string cells.

    Merged cell regions are handled by openpyxl automatically -- merged cells
    return None for all but the top-left cell.  We convert those to empty
    strings for clean display.
    """
    rows: list[list[str]] = []
    for row in ws.iter_rows():
        cells = [_format_cell(cell.value) for cell in row]
        rows.append(cells)
    return rows


def _read_pyxlsb_sheet(sheet: Any) -> list[list[str]]:
    """Read all rows from a pyxlsb sheet object, returning string cells.

    pyxlsb yields rows as lists of Cell objects.  We extract the value from
    each and format it as a string.
    """
    rows: list[list[str]] = []
    for row in sheet.rows():
        cells = [_format_cell(cell.v) for cell in row]
        rows.append(cells)
    return rows


def _format_cell(value: Any) -> str:
    """Convert a cell value to a clean display string.

    Handles None, numbers, booleans, and strings.  Floats that are actually
    integers (e.g. 10.0) are displayed without the decimal point.
    """
    if value is None:
        return ""

    if isinstance(value, bool):
        return "Yes" if value else "No"

    if isinstance(value, float):
        if value == int(value) and not (value != value):  # NaN check
            return str(int(value))
        # Format with reasonable precision -- avoid excessive decimals
        # from floating point representation.
        formatted = f"{value:.6f}".rstrip("0").rstrip(".")
        return formatted

    if isinstance(value, int):
        return str(value)

    # For strings, strip whitespace and replace newlines with spaces
    # to keep the markdown table structure intact.
    text = str(value).strip()
    text = text.replace("\n", " ").replace("\r", " ")
    return text


def _count_data(rows: list[list[str]]) -> tuple[int, int]:
    """Count non-empty rows and total cells with data.

    Returns (data_rows, cells_with_data).  A row is counted as a data row
    if it contains at least one non-empty cell.
    """
    data_rows = 0
    cells_with_data = 0
    for row in rows:
        row_has_data = False
        for cell in row:
            if cell:
                cells_with_data += 1
                row_has_data = True
        if row_has_data:
            data_rows += 1
    return data_rows, cells_with_data


def _rows_to_markdown_table(rows: list[list[str]]) -> str:
    """Convert a list of string rows into a markdown table.

    The first non-empty row is used as the header.  If all rows are data
    (no obvious header), the first row is still treated as the header
    since we have no way to distinguish.

    Empty rows at the beginning and end are trimmed.  Completely empty
    rows in the middle are skipped.
    """
    if not rows:
        return ""

    # Trim leading and trailing empty rows.
    trimmed = _trim_empty_rows(rows)
    if not trimmed:
        return ""

    # Ensure all rows have the same number of columns.
    max_cols = max(len(r) for r in trimmed)
    if max_cols == 0:
        return ""

    normalized: list[list[str]] = []
    for row in trimmed:
        # Skip fully empty rows in the middle.
        if not any(cell for cell in row):
            continue
        padded = list(row)
        while len(padded) < max_cols:
            padded.append("")
        # Truncate extra columns if any.
        normalized.append(padded[:max_cols])

    if not normalized:
        return ""

    # Escape pipe characters in cells.
    escaped: list[list[str]] = []
    for row in normalized:
        escaped.append([cell.replace("|", "\\|") for cell in row])

    header = escaped[0]
    separator = ["---"] * max_cols
    body = escaped[1:]

    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(separator) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(row) + " |")

    return "\n".join(lines)


def _trim_empty_rows(rows: list[list[str]]) -> list[list[str]]:
    """Remove leading and trailing rows where all cells are empty."""
    # Find first non-empty row.
    start = 0
    for i, row in enumerate(rows):
        if any(cell for cell in row):
            start = i
            break
    else:
        return []

    # Find last non-empty row.
    end = len(rows) - 1
    for i in range(len(rows) - 1, -1, -1):
        if any(cell for cell in rows[i]):
            end = i
            break

    return rows[start : end + 1]
